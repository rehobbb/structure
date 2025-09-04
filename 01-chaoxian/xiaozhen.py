import re,pandas as pd
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter
#  ***************************    
def find_chunk(indicator,endflag,file_list):
    chunk = []
    i = 0 
    find_target = False
    end_count = 0
    while i<len(file_list):
        if indicator in file_list[i]:
            find_target = True
            chunk.append(file_list[i])
            i += 1
            continue
        if find_target:
            if endflag == '**':
                if endflag in file_list[i]:
                    end_count += 1
                if end_count == 2:
                    break
                else :
                    chunk.append(file_list[i])
                    i += 1
                    continue
            else:
                if endflag in file_list[i]:
                    break
                else :
                    chunk.append(file_list[i])
                    i += 1
                    continue
        i += 1
    return chunk
#  ***************************    
def find_chunks(indicators,endflag,file_list):
    chunks = []
    for indicator in indicators.values():
        chunk = find_chunk(indicator,endflag,file_list)
        chunks.append(chunk)
    return chunks
#  ***************************    
def is_contained(target,string_list):
    return any(target in s for s in string_list)
#  ***************************    
def get_disp_data(onechunk,index,data_a,indi_key,f_ratio):
    i = 0
    while i<len(onechunk):
        parts1 = onechunk[i].split()
        if parts1[0].isdigit() and int(parts1[0])<100:
            floor = (int(parts1[0])) 
            if is_contained(f_ratio,onechunk):
                parts2 = onechunk[i+1].split()
                value = max(float(parts1[index+2]),float(parts2[index]))
            else:
                 parts2 = re.split(r'\s{2,}',onechunk[i+1].strip())
                 if 'w' in indi_key:
                    value =parts2[index+1].replace('/ ','/')
                 else:
                    value =parts2[index].replace('/ ','/')
            exist_dict = next((item for item in data_a if item['floor'] == floor), None)
            if exist_dict:
                exist_dict[indi_key] = value
            else:
                dict = {'floor':floor,indi_key:value}
                data_a.append(dict)
            i += 1
            continue
        i += 1
#  ***************************       
def get_eforce_data(onechunk,e_index,data_a,indi_key):
     i = 0
     v_pattern = r'(\d+\.?\d*)\(\s*(\d+\.?\d*)%\)'
     while i<len(onechunk):
        parts = re.split(r'\s{2,}',onechunk[i].strip())
        if parts[0].isdigit():
            floor = int(parts[0])
            value_v = float(re.search(v_pattern,parts[e_index]).group(1))
            value_vmr = float(re.search(v_pattern,parts[e_index]).group(2))/100
            value_m = float(parts[e_index+1])
            exist_dict = next((item for item in data_a if item['floor'] == floor),None)
            if exist_dict:
                exist_dict[indi_key+'_v'] = round(value_v,0)
                exist_dict[indi_key+'_vmr'] = round(value_vmr,4)
                exist_dict[indi_key+'_m'] = round(value_m,0)
            else:
                dict = {'floor':floor,indi_key+'_v':round(value_v,0),
                        indi_key+'_vmr':round(value_vmr,4),indi_key+'_m':round(value_m,0)}
                data_a.append(dict)
            i += 1
            continue
        i += 1
#  ***************************    
def get_wforce_data(onechunk,w_index,data_a):
    i = 0
    while i <len(onechunk):
        parts = onechunk[i].split()
        if parts[0].isdigit():
            parts1 = onechunk[i+1].split()
            floor = int(parts[0])
            value_vx = float(parts[w_index])
            value_mx = float(parts[w_index+1])
            value_vy = float(parts1[w_index-2])
            value_my = float(parts1[w_index-1])
            exist_dict = next((item for item in data_a if item['floor'] == floor),None)
            if exist_dict:
                exist_dict['wx_v'] = round(value_vx,0)
                exist_dict['wx_m'] = round(value_mx,0)
                exist_dict['wy_v'] = round(value_vy,0)
                exist_dict['wy_m'] = round(value_my,0)
            else:
                dict = {'floor':floor,'wx_v':round(value_vx,0),'wx_m':round(value_mx,0),
                'wy_v':round(value_vy,0),'wy_m':round(value_my,0)}
                data_a.append(dict)
            i += 1
            continue
        i += 1
#  ***************************  
def get_ratiov_data(onechunk,ratiov_index,data_a):
    i = 0
    while i <len(onechunk):
        parts = onechunk[i].split()
        if parts[0].isdigit():
            floor = int(parts[0])
            ratio_vx = float(parts[ratiov_index])
            ratio_vy = float(parts[ratiov_index+1])
            exist_dict = next((item for item in data_a if item['floor'] == floor),None)
            if exist_dict:
                exist_dict['ratio_vx'] = round(ratio_vx,2)
                exist_dict['ratio_vy'] = round(ratio_vy,2)
            else:
                dict = {'floor':floor,'ratio_vx':round(ratio_vx,2),'ratio_vy':round(ratio_vy,2)}
                data_a.append(dict)
            i += 1
            continue
        i += 1
# ********************* 
def get_ratios_data(onechunk,s_structure,data_a):
    i = 0
    floor_chunks = []
    if '剪' in s_structure:
        rs_x = 'Ratx2'
    else:
        rs_x = 'Ratx1'
    pattern = r'\d+.\d+'
    while i < len(onechunk):
        if 'Floor' in onechunk[i]:
            result = re.search(r'\d+',onechunk[i])
            if result:
                floor = int(result.group())
                floor_chunk = find_chunk('Floor','--',onechunk[i:])
                floor_chunks.append(floor_chunk)
            i += 1
            continue
        i += 1
    for floor_chunk in floor_chunks:  
        j = 0
        floor = re.search(r'\d+',floor_chunk[0]).group()
        while j <len(floor_chunk):           
            if rs_x in floor_chunk[j]:
                ratio_x,ratio_y = map(float,re.findall(pattern,floor_chunk[j]))
                exist_dict = next((item for item in data_a if item['floor'] == int(floor)),None)
                if exist_dict:
                    exist_dict['ratio_sx'] = round(ratio_x,2)
                    exist_dict['ratio_sy'] = round(ratio_y,2)
                else:  
                    dict = {'floor':int(floor),'ratio_sx':round(ratio_x,2),'ratio_y':round(ratio_sy,2)}
                    data_a.append(dict)
                j += 1
                continue
            j += 1
#  ***************************    
direct = input('pleas input the directory:')
direct_wdisp = direct + '\\wdisp.out'
direct_wzq = direct + '\\wzq.out'
direct_wmass = direct + '\\wmass.out'
s_structure = ''
d_index = 3
e_index = 3
w_index = 4
ratiov_index = 4
f_ratio = '规定'
indicators_disp = {
    'ex_disp':'X 方向地震作用下的楼层最大位移',
    'ey_disp':'Y 方向地震作用下的楼层最大位移' ,
    'wx_disp':'+X 方向风荷载作用下的楼层最大位移',
    'wy_disp':'+Y 方向风荷载作用下的楼层最大位移',
    'ex+_dratio':'X+ 偶然偏心规定水平力作用下的楼层最大位移',
    'ex-_dratio':'X- 偶然偏心规定水平力作用下的楼层最大位移',
    'ey+_dratio':'Y+ 偶然偏心规定水平力作用下的楼层最大位移',
    'ey-_dratio':'Y- 偶然偏心规定水平力作用下的楼层最大位移',
           }
indicators_force_e = {
    'ex':'各层 X 方向的作用力(CQC)',
    'ey':'各层 Y 方向的作用力(CQC)',
}
indicator_force_w = '                           风荷载信息'
indicator_ratio_v = '楼层抗剪承载力验算'
indicator_ratio_s = '各层刚心、偏心率、相邻层侧移刚度比等计算信息'
endflag_disp = '=='
endflag_force_e = '='
endflag_force_w = '各楼层等效尺寸'
endflag_ratio_v = '**'
endflag_ratio_s = '**'
data_a = []
indi_keys_disp = list(indicators_disp.keys())
indi_keys_force_e = list(indicators_force_e.keys())
with open(direct_wdisp,'r') as file_d:
    allfile_wdisp = file_d.readlines()
    datachunks_disp = find_chunks(indicators_disp,endflag_disp,allfile_wdisp)
    f_datachunks_disp = [ [s for s in tmpchunk if s.strip()]
    for tmpchunk in datachunks_disp]
    for onechunk,indi_key in zip(f_datachunks_disp ,indi_keys_disp):
        get_disp_data(onechunk,d_index,data_a,indi_key,f_ratio)
with open(direct_wzq,'r') as file_e:
    allfile_wzq = file_e.readlines()
    datachunks_e_force = find_chunks(indicators_force_e,endflag_force_e,allfile_wzq)
    f_datachunks_force_e = [[s for s in tmpchunk if s.strip()]
    for tmpchunk in datachunks_e_force]
    for onechunk,indi_key in zip(f_datachunks_force_e ,indi_keys_force_e):
         get_eforce_data(onechunk,e_index,data_a,indi_key)
with open(direct_wmass,'r') as file_m:
    allfile_wmass = file_m.readlines()
    for line in allfile_wmass:
        if '结构总体信息' in line:
            s_tmp = allfile_wmass[allfile_wmass.index(line)+1]
            s_structure = re.search(r'\s*\w+:\s*(\w+)',s_tmp).group(1)
            break
    datachunk_w_force = find_chunk(indicator_force_w,endflag_force_w,allfile_wmass) 
    f_datachunk_force_w = [s for s in datachunk_w_force if s.strip()]
    get_wforce_data(f_datachunk_force_w,w_index,data_a)
    datachunk_ratio_v = find_chunk(indicator_ratio_v,endflag_ratio_v,allfile_wmass) 
    f_datachunk_ratio_v = [s for s in datachunk_ratio_v if s.strip()]
    get_ratiov_data(f_datachunk_ratio_v,ratiov_index,data_a)
    datachunk_ratio_s = find_chunk(indicator_ratio_s,endflag_ratio_s,allfile_wmass) 
    f_datachunk_ratio_s = [s for s in datachunk_ratio_s if s.strip()]
    get_ratios_data(f_datachunk_ratio_s,s_structure,data_a)
with pd.ExcelWriter(direct + '\\data3.xlsx',engine = 'openpyxl') as writer:
    df = pd.DataFrame(data_a)
    df.to_excel(writer,index = False,sheet_name='YJK')
    worksheet = writer.sheets['YJK']
    for col in worksheet.iter_cols():
        max_length = 0
        for cell in col:
            cell.alignment = Alignment(horizontal='center',vertical='center')
            if len(str(cell.value)) > max_length:
                max_length = len(str(cell.value))
        column_letter = get_column_letter(col[0].column)
        worksheet.column_dimensions[column_letter].width = max_length + 2







